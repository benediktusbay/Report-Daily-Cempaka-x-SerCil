import os
import re
import math
import base64
import hashlib
import io
import json
import secrets
import urllib.error
import urllib.request
from sqlalchemy import inspect, text
from datetime import datetime, timedelta
import datetime as dt
from functools import wraps

import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL', 'sqlite:///' + os.path.join(BASE_DIR, 'sales.db')
).replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 30 * 1024 * 1024
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

db = SQLAlchemy(app)

# Billing mapping agreed for the dashboard.
DEVICE_GROUPS = {'mobile phones', 'tablet'}
MAC_GROUPS = {'computer'}
ACC_GROUPS = {'audio', 'computer accessories', 'mobile accessories', 'tablets accessories', 'wearable'}
QVO_THRESHOLD = 46_000_000
WEEK_PCTS = {1: 0.75, 2: 0.90, 3: 1.00, 4: 1.00}
WEEK_END_DAY = {1: 7, 2: 14, 3: 21, 4: 31}
WEEK_START_DAY = {1: 1, 2: 8, 3: 15, 4: 22}
SKU_TARGETS = {'2-3': 13, '4-6': 6, '7-10': 5, '>10': 2}

# Public Viewer is intentionally limited to these depots.
# Admin continues to see every depot available in Monthly Target / Billing.
VIEWER_ALLOWED_DEPOS = ['Cempaka', 'Serang', 'Cilegon']

# Only these email addresses can request an OTP and open the Stock page.
STOCK_ALLOWED_EMAILS = {
    'fanny.lolowang@erajaya.com',
    'isroudin.01@erajaya.com',
    'rafhyski.alhasan@erajaya.com',
    'zefanya.simorangkir@erajaya.com',
    'ikmah.novtianingrum@erajaya.com',
    'benediktus.kristianto@erajaya.com',
}
STOCK_SESSION_MINUTES = 15
STOCK_DEPOTS = [
    ('Cempaka', 'TAM DC CEMPAKA MAS'),
    ('Cilegon', 'TAM DC CILEGON'),
    ('Roxy', 'TAM DC ROXY'),
    ('Serang', 'TAM DC SERANG'),
    ('Tangerang', 'TAM DC TANGERANG'),
]

# The dashboard is intentionally limited to these eight Apple salesmen.
# Keep this sequence as the canonical display/filter order.
LOCKED_SALESMEN = [
    'Andy Varandy',
    'Michael Serafin Sidik',
    'Muhamad Fajri',
    'Edi Suyitno',
    'Zefanya Septania Simorangkir',
    'Edi Purnomo',
    'Rafhyski Alhasan',
    'Ikmah Novtianingrum',
]

LOYALTY_TARGETS = {
    'CROWN': 5_000_000_000,
    'DIAMOND': 2_000_000_000,
    'GOLD': 500_000_000,
    'SILVER': 350_000_000,
    'BRONZE': 50_000_000,
}

# Program Loyalty master supplied by the business team for Cempaka.
# Dealer/depo/salesman labels are resolved from Monthly Target by BP so an
# updated target master remains the single source for ownership metadata.
LOYALTY_CEMPAKA_BP = {
    '10028000': 'CROWN', '10072814': 'DIAMOND', '10028005': 'CROWN',
    '10002175': 'CROWN', '10002353': 'GOLD', '10005878': 'GOLD',
    '10082147': 'SILVER', '10045600': 'BRONZE', '10003006': 'BRONZE',
    '10066523': 'GOLD', '10044035': 'BRONZE', '10056785': 'BRONZE',
    '10080178': 'SILVER', '10006541': 'BRONZE', '10004520': 'BRONZE',
    '10073725': 'BRONZE', '10054986': 'SILVER', '10000677': 'BRONZE',
    '10005697': 'BRONZE', '10038759': 'GOLD', '10028075': 'BRONZE',
    '10027998': 'DIAMOND', '10034424': 'GOLD', '10034388': 'GOLD',
}


# Incentive organization structure.
INCENTIVE_ORG = {
    'SC': {
        'Zefanya Septania Simorangkir': ['Zefanya Septania Simorangkir'],
        'Rafhyski Alhasan': ['Rafhyski Alhasan'],
        # Ikmah's incentive combines Serang + Cilegon because both depots belong
        # to the same SC coverage.
        'Ikmah Novtianingrum': ['Ikmah Novtianingrum'],
        'Andy Varandy': ['Andy Varandy'],
        'Michael Serafin Sidik': ['Michael Serafin Sidik'],
        'Muhamad Fajri': ['Muhamad Fajri'],
        'Edi Suyitno': ['Edi Suyitno'],
        'Edi Purnomo': ['Edi Purnomo'],
    },
    'ASH': {
        'Fanny Anggraeni Lolowang': [
            'Zefanya Septania Simorangkir',
            'Rafhyski Alhasan',
        ],
        'Isroudin': [
            'Ikmah Novtianingrum',
        ],
        'Andika Polindira': [
            'Andy Varandy',
            'Michael Serafin Sidik',
        ],
        'Miyarni': [
            'Muhamad Fajri',
            'Edi Suyitno',
            'Edi Purnomo',
        ],
    },
    'TSH': {
        'Benediktus Bayu Dwi Kristianto': [
            'Zefanya Septania Simorangkir',
            'Rafhyski Alhasan',
            'Ikmah Novtianingrum',
        ],
        'Frenky Sidarta Hidayat': [
            'Andy Varandy',
            'Michael Serafin Sidik',
            'Muhamad Fajri',
            'Edi Suyitno',
            'Edi Purnomo',
        ],
    },
    'LOB': {
        'Aditya Saputra': [
            'Zefanya Septania Simorangkir',
            'Rafhyski Alhasan',
            'Ikmah Novtianingrum',
            'Andy Varandy',
            'Michael Serafin Sidik',
            'Muhamad Fajri',
            'Edi Suyitno',
            'Edi Purnomo',
        ],
    },
}

INCENTIVE_SCHEME = {
    'SC': {
        'max': 5_000_000,
        'speed_each': 312_500,
        'sku_each': 250_000,
        'qvo': 1_250_000,
        'revenue': {'Device': 1_050_000, 'Macbook': 150_000, 'ACC': 300_000},
    },
    'ASH': {
        'max': 7_000_000,
        'speed_each': 437_500,
        'sku_each': 350_000,
        'qvo': 2_100_000,
        'revenue': {'Device': 1_225_000, 'Macbook': 175_000, 'ACC': 350_000},
    },
    'TSH': {
        'max': 8_500_000,
        'speed_each': 531_250,
        'sku_each': 425_000,
        'qvo': 2_975_000,
        'revenue': {'Device': 1_190_000, 'Macbook': 170_000, 'ACC': 340_000},
    },
    'LOB': {
        'max': 12_000_000,
        'speed_each': 600_000,
        'sku_each': 750_000,
        'qvo': 4_200_000,
        # LOB Apple revenue incentive does not have a separate Mac parameter.
        'revenue': {'Device': 1_800_000, 'ACC': 600_000},
    },
}

# SKU penetration target per SC. Higher levels aggregate by number of SCs covered.
INCENTIVE_SKU_TARGET_PER_SC = {'2-3': 13, '4-6': 6, '7-10': 5, '>10': 2}


# Indonesia national public holidays for 2026.
# Time Gone rule requested: Sunday and national public holidays are not working days.
# Cuti bersama is intentionally NOT excluded unless you later decide to treat it as a non-working day.
NATIONAL_HOLIDAYS_2026 = {
    '2026-01-01',  # New Year
    '2026-01-16',  # Isra Mikraj
    '2026-02-17',  # Chinese New Year
    '2026-03-19',  # Nyepi
    '2026-03-21',  # Idul Fitri
    '2026-03-22',  # Idul Fitri (Sunday anyway)
    '2026-04-03',  # Good Friday
    '2026-04-05',  # Easter (Sunday anyway)
    '2026-05-01',  # Labour Day
    '2026-05-14',  # Ascension Day
    '2026-05-27',  # Idul Adha
    '2026-05-31',  # Vesak (Sunday anyway)
    '2026-06-01',  # Pancasila Day
    '2026-06-16',  # Islamic New Year
    '2026-08-17',  # Independence Day
    '2026-08-25',  # Prophet Muhammad's Birthday
    '2026-12-25',  # Christmas
}


def is_working_day(day):
    """Working day = Monday-Saturday, excluding national public holidays."""
    if day.weekday() == 6:  # Sunday
        return False
    return day.isoformat() not in NATIONAL_HOLIDAYS_2026


def working_day_progress(start, end, as_of):
    """
    Returns elapsed working days, total working days, and Time Gone %.
    as_of is normally latest billing date in the active filter scope.
    """
    total = 0
    elapsed = 0
    cur = start
    capped = min(max(as_of or start, start), end)

    while cur <= end:
        if is_working_day(cur):
            total += 1
            if cur <= capped:
                elapsed += 1
        cur += dt.timedelta(days=1)

    pct = (elapsed / total * 100) if total else 0
    return elapsed, total, pct

# Canonical display names and aliases from Billing Detail.
SALESMAN_ALIASES = {
    'zefanya septania simorangkir': 'Zefanya Septania Simorangkir',
    'rafhyski alhasan': 'Rafhyski Alhasan',
    'ikmah novtianingrum': 'Ikmah Novtianingrum',
    'michael serafin sidik': 'Michael Serafin Sidik',
    'michael serafin sidik deactive': 'Michael Serafin Sidik',
    'andy varandy': 'Andy Varandy',
    'andy varandy deactive': 'Andy Varandy',
    'edi purnomo': 'Edi Purnomo',
    'edi suyitno': 'Edi Suyitno',
    'muhamad fajri': 'Muhamad Fajri',
    'muhamad fajri deactive': 'Muhamad Fajri',
}

# Fallback only when a BP is not found in Monthly Target. Ikmah intentionally has no
# fallback because Serang/Cilegon must be distinguished by BP from the monthly target.
SALESMAN_DEPO_FALLBACK = {
    'Zefanya Septania Simorangkir': 'Cempaka',
    'Rafhyski Alhasan': 'Cempaka',
    'Michael Serafin Sidik': 'Roxy',
    'Andy Varandy': 'Roxy',
    'Edi Purnomo': 'Tangerang',
    'Edi Suyitno': 'Tangerang',
    'Muhamad Fajri': 'Tangerang',
}

COLOR_WORDS = {
    'black','white','blue','green','red','yellow','purple','pink','orange','gray','grey','silver','gold','starlight',
    'midnight','natural','titanium','desert','graphite','space','rose','coral','teal','ultramarine','indigo','lavender',
    'navy','beige','brown','cream','mint','cyan','magenta','violet','jetblack','cosmic','sky','light','dark',
    'sage','mist','deep','stone'
}

# Colour abbreviations that appear in Billing Detail article descriptions.
TABLET_COLOR_CODES = {'SB', 'SIL'}
MAC_COLOR_CODES = {'MDN', 'STL', 'SLV', 'SKY', 'SB', 'SL', 'BLS', 'CIT', 'IND'}
WATCH_COLOR_CODES = {'JB', 'SG', 'ST', 'MI', 'BK'}

# Multi-word marketing colour names found in accessories.
ACC_COLOR_PHRASES = {
    'RAPID RED',
    'BOLT BLACK',
    'SURGE STONE',
    'NITRO NAVY',
}


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='viewer')


# Legacy manual target table is retained so upgrading does not break an existing DB.
class Target(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(7), nullable=False)
    salesman = db.Column(db.String(160), nullable=False)
    device_target = db.Column(db.Float, default=0)
    macbook_target = db.Column(db.Float, default=0)
    acc_target = db.Column(db.Float, default=0)
    bo_target = db.Column(db.Integer, default=25)
    __table_args__ = (db.UniqueConstraint('month','salesman', name='uq_target_month_salesman'),)


class MonthlyTarget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(7), nullable=False, index=True)
    depo = db.Column(db.String(80), nullable=False, index=True)
    bp = db.Column(db.String(80), nullable=False, index=True)
    dealer = db.Column(db.String(255), nullable=False)
    salesman = db.Column(db.String(160), nullable=False, index=True)
    device_target = db.Column(db.Float, default=0)
    macbook_target = db.Column(db.Float, default=0)
    acc_target = db.Column(db.Float, default=0)
    bo_target = db.Column(db.Integer, default=1)
    qvo_target = db.Column(db.Integer, default=1)
    __table_args__ = (db.UniqueConstraint('month','bp', name='uq_monthly_target_month_bp'),)


class UploadLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    uploaded_by = db.Column(db.String(80))
    rows_read = db.Column(db.Integer, default=0)
    rows_added = db.Column(db.Integer, default=0)


class TargetUploadLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(7), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    uploaded_by = db.Column(db.String(80))
    rows_read = db.Column(db.Integer, default=0)
    dealers_loaded = db.Column(db.Integer, default=0)


class Billing(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    row_hash = db.Column(db.String(64), unique=True, nullable=False)
    billing_date = db.Column(db.Date, nullable=False, index=True)
    salesman = db.Column(db.String(160), nullable=False, index=True)
    sold_to_code = db.Column(db.String(80), nullable=False, index=True)
    sold_to_name = db.Column(db.String(255), nullable=False)
    item_group = db.Column(db.String(160), nullable=False)
    article = db.Column(db.String(500), nullable=False)
    quantity = db.Column(db.Float, default=0)
    nett_amount = db.Column(db.Float, default=0)
    nett_amount_with_tax = db.Column(db.Float, default=0)
    category = db.Column(db.String(20), nullable=False)
    sku_key = db.Column(db.String(300))


class StockSnapshot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_date = db.Column(db.Date, nullable=False, unique=True, index=True)
    source_name = db.Column(db.String(255), nullable=False)
    uploaded_by = db.Column(db.String(160), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    source_rows = db.Column(db.Integer, default=0)
    app_rows = db.Column(db.Integer, default=0)


class StockItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    snapshot_id = db.Column(db.Integer, db.ForeignKey('stock_snapshot.id'), nullable=False, index=True)
    material = db.Column(db.String(80), nullable=False, index=True)
    description = db.Column(db.String(500), nullable=False, index=True)
    depot = db.Column(db.String(40), nullable=False, index=True)
    quantity = db.Column(db.Float, default=0)
    __table_args__ = (
        db.UniqueConstraint('snapshot_id', 'material', 'depot', name='uq_stock_snapshot_material_depot'),
    )


class PricelistItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(30), nullable=False, index=True)
    model = db.Column(db.String(500), nullable=False, index=True)
    srp_promo = db.Column(db.Float, default=0)
    stp_promo = db.Column(db.Float, default=0)
    period = db.Column(db.String(100), nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_by = db.Column(db.String(160))
    __table_args__ = (
        db.UniqueConstraint('category', 'model', name='uq_pricelist_category_model'),
    )


class PricelistUploadLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(30), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    rows_loaded = db.Column(db.Integer, default=0)
    uploaded_by = db.Column(db.String(160))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = db.session.get(User, session['user_id'])
        if not user or user.role != 'admin':
            flash('Menu ini hanya untuk Admin.', 'danger')
            return redirect(url_for('dashboard'))
        return fn(*args, **kwargs)
    return wrapper


def stock_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user_id = session.get('user_id')
        if user_id:
            user = db.session.get(User, user_id)
            if user and user.role == 'admin':
                return fn(*args, **kwargs)
        if not stock_session_is_valid():
            clear_stock_session()
            return redirect(url_for('stock_login', next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def clear_stock_session():
    for key in ('stock_email', 'stock_session_expires_at'):
        session.pop(key, None)


def stock_session_is_valid():
    email = normalize_text(session.get('stock_email')).lower()
    expires_raw = session.get('stock_session_expires_at')
    if email not in STOCK_ALLOWED_EMAILS or not expires_raw:
        return False
    try:
        return datetime.utcnow() < datetime.fromisoformat(expires_raw)
    except (TypeError, ValueError):
        return False


def stock_actor():
    user_id = session.get('user_id')
    if user_id:
        user = db.session.get(User, user_id)
        if user and user.role == 'admin':
            return f'Admin: {user.username}'
    return normalize_text(session.get('stock_email')).lower()


def is_stock_admin():
    user_id = session.get('user_id')
    user = db.session.get(User, user_id) if user_id else None
    return bool(user and user.role == 'admin')


def stock_admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_stock_admin():
            flash('Perbarui stock harian hanya dapat dilakukan oleh Admin.', 'danger')
            return redirect(url_for('stock'))
        return fn(*args, **kwargs)
    return wrapper


def normalize_col(s):
    return re.sub(r'[^a-z0-9]+', ' ', str(s).strip().lower()).strip()


def normalize_text(s):
    return re.sub(r'\s+', ' ', str(s or '').strip())


def canonical_salesman(name):
    clean = normalize_text(name)
    key = clean.lower()
    if key in ('', 'nan', 'none'):
        return ''
    if key in SALESMAN_ALIASES:
        return SALESMAN_ALIASES[key]
    return clean.title()


def canonical_depo(value):
    raw = normalize_text(value)
    key = raw.lower()
    for depo in ('Cempaka', 'Serang', 'Cilegon', 'Roxy', 'Tangerang'):
        if depo.lower() in key:
            return depo
    return raw.title() if raw else 'Unmapped'


def normalize_bp(value):
    if value is None:
        return ''
    s = str(value).strip()
    if s.lower() in ('nan', 'none', ''):
        return ''
    # Excel often exposes BP as 10046783.0; BP must be displayed without .0.
    if re.fullmatch(r'\d+\.0+', s):
        s = s.split('.')[0]
    # Scientific notation fallback for numeric BP cells.
    try:
        if re.fullmatch(r'\d+(?:\.\d+)?[eE][+-]?\d+', s):
            s = format(float(s), '.0f')
    except Exception:
        pass
    return s


BILLING_ALIASES = {
    'billing_date': ['billing date','billingdate','bill date'],
    'billing_document': ['billing document', 'billing document number', 'billing doc'],
    'bill_item_no': ['bill item no', 'billing item no', 'bill item number'],
    'salesman': ['salesman name','salesman','sales person','salesperson name'],
    'sold_to_code': ['sold to party code','sold-to party code','sold to code'],
    'sold_to_name': ['sold to party name','sold-to party name','sold to name'],
    'item_group': ['item group desc','item group description','item group'],
    'article': ['article description','article desc','article'],
    'quantity': ['quantity','qty'],
    # Achievement must use the pre-tax value from Billing Detail.
    'nett_amount': [
        'total net amount no tax', 'total nett amount no tax',
        'net amount no tax', 'nett amount no tax',
        'total net amount without tax', 'total nett amount without tax',
        'net value', 'nett value', 'total net value', 'total nett value',
        'amount', 'total amount', 'net amount', 'nett amount',
    ],
    'nett_amount_with_tax': [
        'total net amount with tax', 'total nett amount with tax',
        'net amount with tax', 'nett amount with tax',
    ],
}

TARGET_ALIASES = {
    'depo': ['depo', 'depot'],
    'bp': ['bp', 'sold to party code', 'sold to code'],
    'dealer': ['dealer name', 'dealer', 'sold to party name'],
    'salesman': ['sc name', 'salesman name', 'salesman'],
    'device_target': ['device total', 'device target', 'target device'],
    'acc_target': ['acc total', 'accessories total', 'acc target', 'target acc'],
    'macbook_target': ['mac total', 'macbook total', 'mac target', 'target mac'],
    'bo_target': ['target bo', 'bo target'],
    'qvo_target': ['target qvo', 'qvo target'],
}


def map_columns(columns, aliases, label='file'):
    normalized = {normalize_col(c): c for c in columns}
    mapped = {}
    for key, choices in aliases.items():
        for alias in choices:
            if normalize_col(alias) in normalized:
                mapped[key] = normalized[normalize_col(alias)]
                break
    missing = [k for k in aliases if k not in mapped]
    if missing:
        raise ValueError(f'Kolom wajib tidak ditemukan di {label}: ' + ', '.join(missing))
    return mapped


def classify(item_group):
    g = normalize_text(item_group).lower()
    if g in DEVICE_GROUPS:
        return 'Device'
    if g in MAC_GROUPS:
        return 'Macbook'
    if g in ACC_GROUPS:
        return 'ACC'
    return 'Other'


def _normalize_storage_token(token):
    """Normalize 512 GB -> 512GB and 1 tb -> 1TB."""
    m = re.fullmatch(r'(\d+(?:\.\d+)?)\s*(GB|TB)', normalize_text(token), flags=re.I)
    if not m:
        return normalize_text(token)
    return f'{m.group(1)}{m.group(2).upper()}'


def _strip_colour_tokens(text, colour_codes=None):
    """Remove colour words/codes while keeping product/model information."""
    colour_codes = {str(x).upper() for x in (colour_codes or set())}
    tokens = re.split(r'\s+', re.sub(r'[/,_]+', ' ', normalize_text(text)))
    cleaned = []

    for token in tokens:
        bare = token.strip('()[]{}.,')
        if bare.lower() in COLOR_WORDS:
            continue
        if bare.upper() in colour_codes:
            continue
        cleaned.append(_normalize_storage_token(token))

    return re.sub(r'\s+', ' ', ' '.join(cleaned)).strip()


def sku_from_article(article, item_group):
    """
    KPI SKU applies to Device + Macbook + ACC.

    Main rule:
      SKU = model/type + storage; colour differences are ignored.

    Category handling:
    - Mobile Phones / Tablet:
      Keep the model/specification and storage, remove colour names/codes.
      This preserves model markers such as M4/M5 that can appear after storage.
    - Computer / Macbook:
      SKU is Mac family + screen/model size + chip + primary storage.
      RAM/GPU and colour variations do not create another SKU.
    - ACC:
      Keep the product/model description and storage when present,
      while removing colour names/codes. For accessories without storage,
      the colour-free model description itself is the SKU.
    """
    group = normalize_text(item_group).lower()
    s = normalize_text(article)

    if not s or s.lower() in ('nan', 'none'):
        return None

    # DEVICE: Mobile Phones + Tablet.
    if group in DEVICE_GROUPS:
        colour_codes = TABLET_COLOR_CODES if group == 'tablet' else set()
        base = _strip_colour_tokens(s, colour_codes)
        return base.upper() if base else None

    # MACBOOK / COMPUTER.
    if group in MAC_GROUPS:
        normalized = re.sub(r'\s+', ' ', re.sub(r'[/,_]+', ' ', s)).strip()
        upper = normalized.upper()

        # In Billing Detail the final GB/TB capacity is the primary SSD/storage.
        capacities = re.findall(r'\b\d+(?:\.\d+)?\s*(?:GB|TB)\b', normalized, flags=re.I)
        storage = _normalize_storage_token(capacities[-1]) if capacities else ''

        # Model family and display/model size.
        if upper.startswith('MB NEO'):
            model_match = re.match(r'\bMB\s+NEO\s+([0-9.]+)', upper)
            model = f"MB NEO {model_match.group(1)}" if model_match else 'MB NEO'
            chip = ''
        else:
            model_match = re.match(r'\b(MBA|MBP)\s+([0-9.]+)', upper)
            model = (
                f"{model_match.group(1)} {model_match.group(2)}"
                if model_match else upper.split()[0]
            )

            # Examples: M3, M5, M5 Pro, M5 Max.
            chip_match = re.search(r'\bM\d+\b(?:\s+(?:PRO|MAX))?', upper)
            chip = chip_match.group(0) if chip_match else ''

        key = ' '.join(x for x in (model, chip, storage) if x)
        if key:
            return re.sub(r'\s+', ' ', key).strip().upper()

        # Defensive fallback.
        fallback = _strip_colour_tokens(s, MAC_COLOR_CODES)
        return fallback.upper() if fallback else None

    # ACCESSORIES: Audio, Computer Accessories, Mobile Accessories,
    # Tablets Accessories, Wearable.
    if group in ACC_GROUPS:
        cleaned = s.upper()

        # Watch S11 SKU follows series + case size. Colour, material and band
        # variants do not create additional SKU records.
        if group == 'wearable':
            watch_s11 = re.search(
                r'\b(?:APP(?:LE)?\s+)?WATCH\s+S?\s*11\s+(42|46)(?:\s*MM)?\b',
                cleaned,
                flags=re.I,
            )
            if watch_s11:
                return f'APP WATCH 11 {watch_s11.group(1)}MM'

        # Remove known multi-word accessory colour names first.
        for phrase in ACC_COLOR_PHRASES:
            cleaned = re.sub(r'\b' + re.escape(phrase) + r'\b', ' ', cleaned, flags=re.I)

        colour_codes = WATCH_COLOR_CODES if group == 'wearable' else set()
        cleaned = _strip_colour_tokens(cleaned, colour_codes)
        return cleaned.upper() if cleaned else None

    return None


def row_hash(vals):
    raw = '|'.join(str(v) for v in vals)
    return hashlib.sha256(raw.encode('utf-8', errors='ignore')).hexdigest()


def month_range(month):
    start = pd.to_datetime(month + '-01').date()
    end = (pd.Timestamp(start) + pd.offsets.MonthEnd(0)).date()
    return start, end


def to_num(value, default=0):
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    text_value = normalize_text(value)
    if not text_value or text_value.lower() in ('nan', 'none', '-'):
        return default
    # Accept common Excel/export formats: Rp 1.234.567, 1.234.567,00,
    # and 1,234,567.00.
    cleaned = re.sub(r'[^0-9,.-]', '', text_value)
    if not cleaned or cleaned in ('-', '.', ','):
        return default
    if ',' in cleaned and '.' in cleaned:
        if cleaned.rfind(',') > cleaned.rfind('.'):
            cleaned = cleaned.replace('.', '').replace(',', '.')
        else:
            cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        tail = cleaned.rsplit(',', 1)[1]
        cleaned = cleaned.replace(',', '.') if len(tail) in (1, 2) else cleaned.replace(',', '')
    elif cleaned.count('.') > 1:
        cleaned = cleaned.replace('.', '')
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return default


def billing_date_value(value):
    """Return a date from an Excel billing cell without creating a DataFrame."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    parsed = pd.to_datetime(value, errors='coerce')
    return None if pd.isna(parsed) else parsed.date()


def iter_billing_xlsx(stream):
    """Stream Billing Detail rows from sheet Export in read-only mode."""
    stream.seek(0)
    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        if 'Export' not in workbook.sheetnames:
            raise ValueError('Sheet Export tidak ditemukan di Billing Detail.')
        worksheet = workbook['Export']
        rows = worksheet.iter_rows(values_only=True)
        column_indexes = None

        # Normally the header is the first row. Searching a few rows also
        # supports exports that contain a title line above the real header.
        for _ in range(30):
            candidate = next(rows, None)
            if candidate is None:
                break
            try:
                mapped = map_columns(candidate, BILLING_ALIASES, 'Billing Detail / sheet Export')
            except ValueError:
                continue
            header = list(candidate)
            column_indexes = {
                key: header.index(source_column)
                for key, source_column in mapped.items()
            }
            break

        if column_indexes is None:
            raise ValueError('Header Billing Detail tidak ditemukan pada sheet Export.')

        for row in rows:
            yield {
                key: row[index] if index < len(row) else None
                for key, index in column_indexes.items()
            }
    finally:
        workbook.close()


def number_id(value):
    try:
        number = float(value or 0)
    except Exception:
        number = 0
    if number.is_integer():
        return f'{int(number):,}'.replace(',', '.')
    return f'{number:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')


def normalize_stock_depot(value):
    raw = normalize_text(value).upper()
    for short_name, source_name in STOCK_DEPOTS:
        if raw == source_name.upper():
            return short_name
    return ''


def stock_column_map(columns, require_material_group=False):
    normalized = {normalize_col(column): column for column in columns}
    aliases = {
        'material': ['material'],
        'description': ['material description', 'material desc', 'description'],
        'depot': ['name 1', 'name1', 'depo', 'depot'],
        'quantity': ['unrestricted', 'stock', 'quantity', 'qty'],
    }
    if require_material_group:
        aliases['material_group'] = ['material group', 'materialgroup']
        aliases['storage_location'] = ['storage location', 'stor location', 'storage loc', 'sloc', 's loc']
    mapped = {}
    for key, choices in aliases.items():
        for choice in choices:
            found = normalized.get(normalize_col(choice))
            if found is not None:
                mapped[key] = found
                break
    missing = [key for key in aliases if key not in mapped]
    if missing:
        raise ValueError('Kolom stok tidak ditemukan: ' + ', '.join(missing))
    return mapped


def save_stock_snapshot(dataframe, stock_date, source_name, uploaded_by, require_material_group=False):
    cmap = stock_column_map(dataframe.columns, require_material_group=require_material_group)
    source_rows = len(dataframe)
    grouped = {}
    app_rows = 0

    for _, row in dataframe.iterrows():
        if require_material_group:
            material_group = normalize_text(row[cmap['material_group']]).upper()
            if not material_group.endswith('APP'):
                continue
            storage_location = normalize_bp(row[cmap['storage_location']])
            if storage_location != '1001':
                continue

        material = normalize_bp(row[cmap['material']])
        description = normalize_text(row[cmap['description']])
        depot = normalize_stock_depot(row[cmap['depot']])
        quantity = to_num(row[cmap['quantity']])
        if not material or not description or not depot:
            continue
        app_rows += 1
        key = (material, depot)
        record = grouped.setdefault(key, {'description': description, 'quantity': 0.0})
        record['quantity'] += quantity

    if not grouped:
        raise ValueError('Tidak ada data stock APP untuk lima Depo yang dapat disimpan.')

    existing = StockSnapshot.query.filter_by(stock_date=stock_date).first()
    if existing:
        StockItem.query.filter_by(snapshot_id=existing.id).delete(synchronize_session=False)
        db.session.delete(existing)
        db.session.flush()

    snapshot = StockSnapshot(
        stock_date=stock_date,
        source_name=source_name,
        uploaded_by=uploaded_by,
        source_rows=source_rows,
        app_rows=app_rows,
    )
    db.session.add(snapshot)
    db.session.flush()
    for (material, depot), record in grouped.items():
        db.session.add(StockItem(
            snapshot_id=snapshot.id,
            material=material,
            description=record['description'],
            depot=depot,
            quantity=record['quantity'],
        ))
    db.session.commit()
    return snapshot, len(grouped)


def send_stock_otp(recipient, code):
    sender = normalize_text(os.environ.get('OTP_SENDER_EMAIL'))
    api_key = normalize_text(os.environ.get('MAILJET_API_KEY'))
    secret_key = normalize_text(os.environ.get('MAILJET_SECRET_KEY'))
    if not sender or not api_key or not secret_key:
        raise RuntimeError('Pengaturan OTP Mailjet di Render belum lengkap.')

    payload = {
        'Messages': [{
            'From': {
                'Email': sender,
                'Name': normalize_text(os.environ.get('OTP_SENDER_NAME')) or 'Stock Apple',
            },
            'To': [{'Email': recipient}],
            'Subject': 'Kode OTP Stock Apple',
            'TextPart': (
                f'Kode OTP Anda: {code}\n\n'
                'Kode berlaku selama 10 menit. Jangan berikan kode ini kepada siapa pun.'
            ),
        }],
    }
    credentials = base64.b64encode(f'{api_key}:{secret_key}'.encode('utf-8')).decode('ascii')
    api_request = urllib.request.Request(
        'https://api.mailjet.com/v3.1/send',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'accept': 'application/json',
            'authorization': f'Basic {credentials}',
            'content-type': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(api_request, timeout=20) as response:
            response_data = json.loads(response.read().decode('utf-8'))
            message_status = response_data.get('Messages', [{}])[0].get('Status')
            if response.status not in (200, 201, 202) or message_status != 'success':
                raise RuntimeError(f'Mailjet belum menerima email OTP (status: {message_status or response.status}).')
    except urllib.error.HTTPError as exc:
        try:
            detail = json.loads(exc.read().decode('utf-8')).get('message', '')
        except Exception:
            detail = ''
        raise RuntimeError(f'Mailjet menolak pengiriman ({exc.code}): {detail or exc.reason}') from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f'Tidak dapat terhubung ke Mailjet: {exc.reason}') from exc


def rupiah(x):
    try:
        return 'Rp {:,.0f}'.format(float(x)).replace(',', '.')
    except Exception:
        return 'Rp 0'


def rupiah_short(x):
    try:
        n = float(x or 0)
    except Exception:
        n = 0
    if abs(n) >= 1_000_000_000:
        txt = f'{n/1_000_000_000:.2f}'.rstrip('0').rstrip('.').replace('.', ',')
        return f'Rp{txt} M'
    if abs(n) >= 1_000_000:
        txt = f'{n/1_000_000:.1f}'.rstrip('0').rstrip('.').replace('.', ',')
        return f'Rp{txt} Jt'
    if abs(n) >= 1_000:
        txt = f'{n/1_000:.1f}'.rstrip('0').rstrip('.').replace('.', ',')
        return f'Rp{txt} Rb'
    return f'Rp{n:,.0f}'.replace(',', '.')


app.jinja_env.filters['rupiah'] = rupiah
app.jinja_env.filters['rupiah_short'] = rupiah_short
app.jinja_env.filters['number_id'] = number_id


@app.before_request
def ensure_db():
    db.create_all()
    # db.create_all() does not add columns to an existing Render database.
    # Apply this small backward-compatible migration automatically.
    billing_columns = {c['name'] for c in inspect(db.engine).get_columns('billing')}
    if 'nett_amount_with_tax' not in billing_columns:
        db.session.execute(text(
            'ALTER TABLE billing ADD COLUMN nett_amount_with_tax FLOAT DEFAULT 0'
        ))
        db.session.commit()
    if User.query.count() == 0:
        username = os.environ.get('ADMIN_USERNAME', 'admin')
        password = os.environ.get('ADMIN_PASSWORD', 'admin123')
        db.session.add(User(username=username, password_hash=generate_password_hash(password), role='admin'))
        db.session.commit()


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'GET' and session.get('user_id') and session.get('role') == 'admin':
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')

        # Username login is case-insensitive:
        # Bayu / bayu / BAYU all resolve to the same account.
        user = User.query.filter(
            db.func.lower(User.username) == username.lower()
        ).first()

        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('dashboard'))
        flash('Username atau password salah.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    # Clear Viewer/Admin session and open the login page.
    # This allows a public Viewer to switch to Admin intentionally.
    session.clear()
    return redirect(url_for('login'))


def target_lookup_for_month(month):
    targets = MonthlyTarget.query.filter_by(month=month).all()
    return targets, {normalize_bp(t.bp): t for t in targets}


def resolve_billing_owner(row, target_by_bp):
    bp = normalize_bp(row.sold_to_code)
    target = target_by_bp.get(bp)

    # Achievement ownership always follows Salesman Name in Billing Detail.
    # Monthly Target supplies dealer/depo metadata only; it must never move a
    # transaction from one salesman to another.
    billing_salesman = canonical_salesman(row.salesman)
    if target:
        return billing_salesman, target.depo, target.dealer, target
    return (
        billing_salesman,
        SALESMAN_DEPO_FALLBACK.get(billing_salesman, 'Unmapped'),
        normalize_text(row.sold_to_name),
        None,
    )


def matches_scope(salesman, depo, salesman_filters, depo_filters):
    if salesman_filters and canonical_salesman(salesman) not in salesman_filters:
        return False
    if depo_filters and depo not in depo_filters:
        return False
    return True


def business_round(value):
    """
    Company rounding rule:
      fractional 0.00 - 0.59 -> round down
      fractional 0.60 - 0.99 -> round up

    Examples: 22.50 -> 22, 18.75 -> 19.
    """
    value = float(value or 0)
    floor_value = math.floor(value)
    fraction = value - floor_value
    return floor_value + (1 if fraction >= 0.60 else 0)


def weekly_targets(bo_target):
    """
    Dashboard Speed Distribution target:
    use round-up (ceil), so with monthly BO target 25:
      Week 1 = 19
      Week 2 = 23
      Week 3 = 25
      Week 4 = 25
    """
    return {
        w: int(math.ceil((bo_target or 0) * WEEK_PCTS[w]))
        for w in range(1, 5)
    }


def incentive_weekly_targets(bo_target):
    """
    Incentive Speed Distribution target:
    keep the agreed company rounding rule:
      0.00-0.59 down, 0.60-0.99 up.

    With monthly BO target 25:
      Cycle 1 = 19
      Cycle 2 = 22
      Cycle 3 = 25
      Cycle 4 = 25
    """
    return {
        w: business_round((bo_target or 0) * WEEK_PCTS[w])
        for w in range(1, 5)
    }


def is_bo_amounts(device_amount=0, macbook_amount=0):
    """
    BO rule:
    - Device saja = BO
    - Macbook saja = BO
    - Device + ACC = BO
    - Macbook + ACC = BO
    - Device + Macbook (+/- ACC) = BO
    - ACC saja = bukan BO

    ACC tidak menjadi syarat BO; dealer qualify selama ada Device dan/atau Macbook.
    """
    return (float(device_amount or 0) + float(macbook_amount or 0)) > 0


def sku_bucket(n):
    if n == 1:
        return '1'
    if 2 <= n <= 3:
        return '2-3'
    if 4 <= n <= 6:
        return '4-6'
    if 7 <= n <= 10:
        return '7-10'
    if n > 10:
        return '>10'
    return '0'


def incentive_status(actual, target, active=True):
    if not active:
        return 'Pending'
    if target and actual >= target:
        return 'Achieved'
    return 'Not Achieved'


def incentive_pct(actual, target):
    return (actual / target * 100) if target else 0


def build_incentive_metrics(month, member_salesmen):
    """
    Aggregate target + actual by the SC names covered by one incentive recipient.
    Dealer-based KPIs use (salesman, BP) as the unique dealer key.

    A BP remains owned by the Salesman Name from Billing Detail, including
    UNMAPPED rows.  Using BP alone would incorrectly merge the same dealer when
    it appears under two different salesmen in one hierarchy calculation.
    """
    start, end = month_range(month)
    member_salesmen = {canonical_salesman(x) for x in member_salesmen}

    monthly_targets, target_by_bp = target_lookup_for_month(month)
    target_rows = [
        t for t in monthly_targets
        if canonical_salesman(t.salesman) in member_salesmen
    ]

    revenue_target = {'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0}
    bo_target = 0
    for t in target_rows:
        revenue_target['Device'] += float(t.device_target or 0)
        revenue_target['Macbook'] += float(t.macbook_target or 0)
        revenue_target['ACC'] += float(t.acc_target or 0)
        bo_target += int(t.bo_target or 0)

    # Defensive fallback: if a target file has no BO rows, preserve the SC monthly
    # default used by the dashboard.
    if not bo_target:
        bo_target = 25 * len(member_salesmen)

    billing_rows = Billing.query.filter(
        Billing.billing_date >= start,
        Billing.billing_date <= end
    ).all()

    dealer = {}
    relevant_billing = []

    # Start with target dealers so BP ownership is stable even with no sales.
    for t in target_rows:
        bp = normalize_bp(t.bp)
        if not bp:
            continue
        target_owner = canonical_salesman(t.salesman)
        dealer[(target_owner, bp)] = {
            'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0,
            'skus': set(),
        }

    for r in billing_rows:
        owner, depo, dealer_name, target = resolve_billing_owner(r, target_by_bp)
        owner = canonical_salesman(owner)
        if owner not in member_salesmen:
            continue

        relevant_billing.append((r, owner, depo))
        bp = normalize_bp(r.sold_to_code)
        if not bp:
            continue
        dealer_key = (owner, bp)
        d = dealer.setdefault(dealer_key, {
            'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0,
            'skus': set(),
        })

        if r.category in ('Device', 'Macbook', 'ACC'):
            d[r.category] += float(r.nett_amount or 0)
            current_sku = sku_from_article(r.article, r.item_group)
            if current_sku:
                d['skus'].add(current_sku)

    revenue_actual = {'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0}
    bo_actual = 0
    qvo_actual = 0
    sku_bins = {'1': 0, '2-3': 0, '4-6': 0, '7-10': 0, '>10': 0}

    for dealer_key, d in dealer.items():
        for cat in revenue_actual:
            revenue_actual[cat] += d[cat]

        total_sales = d['Device'] + d['Macbook'] + d['ACC']
        if is_bo_amounts(d['Device'], d['Macbook']):
            bo_actual += 1
        if total_sales >= QVO_THRESHOLD:
            qvo_actual += 1

        bucket = sku_bucket(len(d['skus']))
        if bucket in sku_bins:
            sku_bins[bucket] += 1

    latest_date = max(
        [r.billing_date for r, _, _ in relevant_billing],
        default=None
    )

    # Speed Distribution: cumulative unique BO by cycle cut-off.
    speed_targets = incentive_weekly_targets(bo_target)
    speed_actual = {}
    speed_active = {}

    for cycle in range(1, 5):
        active = bool(latest_date and latest_date.day >= WEEK_START_DAY[cycle])
        speed_active[cycle] = active

        if not active:
            speed_actual[cycle] = 0
            continue

        cutoff_day = min(WEEK_END_DAY[cycle], end.day, latest_date.day)
        cutoff = start.replace(day=cutoff_day)
        bo_by_owner_bp = {}

        for r, owner, depo in relevant_billing:
            if r.billing_date > cutoff:
                continue
            bp = normalize_bp(r.sold_to_code)
            if not bp:
                continue
            vals = bo_by_owner_bp.setdefault(
                (owner, bp),
                {'Device': 0.0, 'Macbook': 0.0},
            )
            if r.category == 'Device':
                vals['Device'] += float(r.nett_amount or 0)
            elif r.category == 'Macbook':
                vals['Macbook'] += float(r.nett_amount or 0)

        speed_actual[cycle] = sum(
            1 for vals in bo_by_owner_bp.values()
            if is_bo_amounts(vals['Device'], vals['Macbook'])
        )

    qvo_target = business_round(bo_target * 0.50)

    sc_count = len(member_salesmen)
    sku_targets = {
        bucket: target * sc_count
        for bucket, target in INCENTIVE_SKU_TARGET_PER_SC.items()
    }

    return {
        'bo_target': bo_target,
        'bo_actual': bo_actual,
        'speed_targets': speed_targets,
        'speed_actual': speed_actual,
        'speed_active': speed_active,
        'qvo_target': qvo_target,
        'qvo_actual': qvo_actual,
        'sku_targets': sku_targets,
        'sku_actual': sku_bins,
        'revenue_target': revenue_target,
        'revenue_actual': revenue_actual,
        'latest_date': latest_date,
        'sc_count': sc_count,
    }


def calculate_incentive(level, person, month):
    level = level.upper()
    scheme = INCENTIVE_SCHEME[level]
    members = INCENTIVE_ORG[level][person]
    metrics = build_incentive_metrics(month, members)

    detail = []
    earned = 0

    # Speed Distribution: 4 independent binary parameters.
    for cycle in range(1, 5):
        target = metrics['speed_targets'][cycle]
        actual = metrics['speed_actual'][cycle]
        active = metrics['speed_active'][cycle]
        status = incentive_status(actual, target, active)
        payout = scheme['speed_each'] if status == 'Achieved' else 0
        earned += payout
        detail.append({
            'kpi': 'Speed Distribution',
            'parameter': f'Cycle {cycle}',
            'target': target,
            'achievement': actual if active else None,
            'pct': incentive_pct(actual, target) if active else None,
            'status': status,
            'payout': payout,
            'format': 'count',
        })

    # Productivity / SKU: 4 independent binary parameters.
    for bucket in ('2-3', '4-6', '7-10', '>10'):
        target = metrics['sku_targets'][bucket]
        actual = metrics['sku_actual'][bucket]
        status = incentive_status(actual, target, True)
        payout = scheme['sku_each'] if status == 'Achieved' else 0
        earned += payout
        detail.append({
            'kpi': 'SKU Penetration',
            'parameter': f'SKU {bucket}',
            'target': target,
            'achievement': actual,
            'pct': incentive_pct(actual, target),
            'status': status,
            'payout': payout,
            'format': 'count',
        })

    # QVO: Apple parameter is 50% of total monthly BO target for every level,
    # including LOB.
    qvo_target = metrics['qvo_target']
    qvo_actual = metrics['qvo_actual']
    qvo_status = incentive_status(qvo_actual, qvo_target, True)
    qvo_payout = scheme['qvo'] if qvo_status == 'Achieved' else 0
    earned += qvo_payout
    detail.append({
        'kpi': 'QVO',
        'parameter': '50% of Total BO Target',
        'target': qvo_target,
        'achievement': qvo_actual,
        'pct': incentive_pct(qvo_actual, qvo_target),
        'status': qvo_status,
        'payout': qvo_payout,
        'format': 'count',
    })

    # Revenue: category parameters differ by level.
    for category, max_payout in scheme['revenue'].items():
        target = metrics['revenue_target'][category]
        actual = metrics['revenue_actual'][category]
        status = incentive_status(actual, target, True)
        payout = max_payout if status == 'Achieved' else 0
        earned += payout
        detail.append({
            'kpi': 'Revenue',
            'parameter': category,
            'target': target,
            'achievement': actual,
            'pct': incentive_pct(actual, target),
            'status': status,
            'payout': payout,
            'format': 'rupiah',
        })

    return {
        'level': level,
        'person': person,
        'members': members,
        'sc_count': metrics['sc_count'],
        'earned': earned,
        'max_incentive': scheme['max'],
        'earned_pct': incentive_pct(earned, scheme['max']),
        'detail': detail,
        'latest_date': metrics['latest_date'],
    }


@app.route('/')
def dashboard():
    # Public Viewer mode:
    # anyone with the dashboard URL may view reports without logging in.
    # Admin-only routes still require a real authenticated admin session.
    if 'user_id' not in session:
        session['username'] = 'Viewer'
        session['role'] = 'viewer'
    latest = db.session.query(db.func.max(Billing.billing_date)).scalar()
    latest_target_month = db.session.query(db.func.max(MonthlyTarget.month)).scalar()
    default_month = latest.strftime('%Y-%m') if latest else (latest_target_month or datetime.now().strftime('%Y-%m'))
    month = request.args.get('month', default_month)

    # Multi-select Depo.
    # IMPORTANT: Viewer restriction is enforced in backend, not only hidden in HTML,
    # so a Viewer cannot expose another depo by editing the URL manually.
    requested_depos = [
        normalize_text(x)
        for x in request.args.getlist('depo')
        if normalize_text(x)
    ]

    if session.get('role') == 'admin':
        # Dashboard awal selalu menampilkan tiga depo operasional utama.
        depo_filters = requested_depos or VIEWER_ALLOWED_DEPOS.copy()
    else:
        allowed = set(VIEWER_ALLOWED_DEPOS)
        depo_filters = [d for d in requested_depos if d in allowed]

        # No valid selection means "All Viewer Depo", not "All Company Depo".
        # Therefore Viewer always remains scoped to Cempaka + Serang + Cilegon.
        if not depo_filters:
            depo_filters = VIEWER_ALLOWED_DEPOS.copy()

    # Multi-select Salesman. No selection means All Salesman.
    requested_salesmen = [
        canonical_salesman(x)
        for x in request.args.getlist('salesman')
        if canonical_salesman(x)
    ]
    start, end = month_range(month)

    monthly_targets, target_by_bp = target_lookup_for_month(month)
    billing_rows = Billing.query.filter(Billing.billing_date >= start, Billing.billing_date <= end).all()

    # Filter dropdown options come from target master.
    all_depos = sorted({t.depo for t in monthly_targets if t.depo and t.depo != 'Unmapped'})

    if session.get('role') == 'admin':
        depos = all_depos
    else:
        # Keep the requested business order for Viewer.
        available = set(all_depos)
        depos = [d for d in VIEWER_ALLOWED_DEPOS if d in available]

    # Full operational scope must not depend on Monthly Target availability.
    # This keeps Billing achievement visible even before the target master
    # for the selected month has been uploaded.
    full_operational_scope = (
        not monthly_targets
        and set(depo_filters) == set(VIEWER_ALLOWED_DEPOS)
    )

    salesman_options = set()
    for t in monthly_targets:
        if not depo_filters or t.depo in depo_filters:
            salesman_options.add(t.salesman)
    for r in billing_rows:
        owner, depo, _, _ = resolve_billing_owner(r, target_by_bp)
        if full_operational_scope:
            if owner in LOCKED_SALESMEN:
                salesman_options.add(owner)
        elif owner and (not depo_filters or depo in depo_filters):
            salesman_options.add(owner)
    # Do not allow unexpected/raw billing names to create extra dashboard rows.
    salesmen = [s for s in LOCKED_SALESMEN if s in salesman_options]

    # Keep only valid selections and preserve the canonical dashboard order.
    requested_set = set(requested_salesmen)
    salesman_filters = [s for s in salesmen if s in requested_set]

    # BO belongs to the salesman, not to a depo mapping. Calculate it from all
    # billing rows in the selected month before applying any depo filter, so a
    # BP remains part of the salesman's BO whether it is mapped or UNMAPPED.
    # The (salesman, BP) key also prevents the same BP from being counted twice.
    bo_billing = []
    bo_amounts_by_owner_bp = {}
    for r in billing_rows:
        owner, _, _, _ = resolve_billing_owner(r, target_by_bp)
        if not owner or (salesman_filters and owner not in salesman_filters):
            continue
        bp = normalize_bp(r.sold_to_code)
        if not bp:
            continue
        bo_billing.append((r, owner))
        amounts = bo_amounts_by_owner_bp.setdefault(
            (owner, bp), {'device': 0.0, 'macbook': 0.0}
        )
        if r.category == 'Device':
            amounts['device'] += float(r.nett_amount or 0)
        elif r.category == 'Macbook':
            amounts['macbook'] += float(r.nett_amount or 0)

    bo_actual_by_salesman = {}
    for (owner, _), amounts in bo_amounts_by_owner_bp.items():
        if is_bo_amounts(amounts['device'], amounts['macbook']):
            bo_actual_by_salesman[owner] = bo_actual_by_salesman.get(owner, 0) + 1

    # Target rows in the active scope.
    scoped_targets = [
        t for t in monthly_targets
        if matches_scope(t.salesman, t.depo, salesman_filters, depo_filters)
    ]

    # Saat ketiga depo operasional utama ditampilkan bersama, achievement harus
    # mengikuti Salesman Name pada Billing Detail. Jangan membuang transaksi
    # hanya karena BP belum ada di Target Master atau metadata depo BP berbeda.
    # Filter depo tetap dipakai saat pengguna memilih sebagian depo supaya
    # pembagian Serang/Cilegon (khususnya untuk Ikmah) tetap akurat.

    # Dealer master starts with all monthly target dealers so zero-achievement dealers remain visible.
    dealer = {}
    for t in scoped_targets:
        bp = normalize_bp(t.bp)
        key = (canonical_salesman(t.salesman), bp)
        dealer[key] = {
            'salesman': t.salesman, 'depo': t.depo, 'bp': bp, 'dealer': t.dealer,
            'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0, 'skus': set(), 'last_date': None,
            'device_items': {},
            'device_target': float(t.device_target or 0), 'macbook_target': float(t.macbook_target or 0),
            'acc_target': float(t.acc_target or 0), 'bo_target': int(t.bo_target or 0), 'qvo_target': int(t.qvo_target or 0),
            'is_target': True
        }

    scoped_billing = []
    for r in billing_rows:
        owner, depo, dealer_name, target = resolve_billing_owner(r, target_by_bp)
        if full_operational_scope:
            billing_matches_scope = (
                bool(owner)
                and owner in LOCKED_SALESMEN
                and (not salesman_filters or owner in salesman_filters)
            )
        else:
            billing_matches_scope = matches_scope(
                owner, depo, salesman_filters, depo_filters
            )
        if not billing_matches_scope:
            continue
        scoped_billing.append((r, owner, depo))
        bp = normalize_bp(r.sold_to_code)
        dealer_key = (canonical_salesman(owner), bp)
        d = dealer.setdefault(dealer_key, {
            'salesman': owner, 'depo': depo, 'bp': bp, 'dealer': dealer_name,
            'Device': 0.0, 'Macbook': 0.0, 'ACC': 0.0, 'skus': set(), 'last_date': None,
            'device_items': {},
            'device_target': float(target.device_target or 0) if target else 0,
            'macbook_target': float(target.macbook_target or 0) if target else 0,
            'acc_target': float(target.acc_target or 0) if target else 0,
            'bo_target': int(target.bo_target or 0) if target else 0,
            'qvo_target': int(target.qvo_target or 0) if target else 0,
            'is_target': bool(target)
        })
        if r.category in ('Device','Macbook','ACC'):
            d[r.category] += float(r.nett_amount or 0)
        if r.category == 'Device':
            item_name = normalize_text(r.article) or '(Tanpa deskripsi)'
            item = d['device_items'].setdefault(item_name, {
                'type': item_name, 'qty': 0.0, 'value': 0.0,
            })
            item['qty'] += float(r.quantity or 0)
            item['value'] += float(r.nett_amount or 0)
        # Recalculate SKU from Article Description on every dashboard load.
        # KPI SKU includes Device + Macbook + ACC. Historical stored sku_key
        # values are intentionally ignored so the latest normalization rule
        # automatically corrects old billing data.
        if r.category in ('Device', 'Macbook', 'ACC'):
            current_sku = sku_from_article(r.article, r.item_group)
            if current_sku:
                d['skus'].add(current_sku)
        if d['last_date'] is None or r.billing_date > d['last_date']:
            d['last_date'] = r.billing_date

    # Salesman target totals from Monthly Target.
    salesman_target = {}
    for t in scoped_targets:
        s = salesman_target.setdefault(t.salesman, {'device':0.0,'macbook':0.0,'acc':0.0,'bo':0,'qvo':0,'dealers':0})
        s['device'] += float(t.device_target or 0)
        s['macbook'] += float(t.macbook_target or 0)
        s['acc'] += float(t.acc_target or 0)
        s['bo'] += int(t.bo_target or 0)
        s['qvo'] += int(t.qvo_target or 0)
        s['dealers'] += 1

    salesman_actual = {}
    sku_detail = []
    dealer_detail = []
    active_dealers = 0
    for dealer_key, d in dealer.items():
        bp = d['bp']
        total = d['Device'] + d['Macbook'] + d['ACC']
        bo = (d['Device'] + d['Macbook']) > 0
        qvo = total >= QVO_THRESHOLD
        sku_count = len(d['skus'])
        bucket = sku_bucket(sku_count)
        if total != 0:
            active_dealers += 1
        s = salesman_actual.setdefault(d['salesman'], {
            'device':0.0,'macbook':0.0,'acc':0.0,'bo':0,'qvo':0,
            'sku_bins':{'1':0,'2-3':0,'4-6':0,'7-10':0,'>10':0}
        })
        s['device'] += d['Device']; s['macbook'] += d['Macbook']; s['acc'] += d['ACC']
        if bo: s['bo'] += 1
        if qvo: s['qvo'] += 1
        if bucket in s['sku_bins']:
            s['sku_bins'][bucket] += 1
        if sku_count > 0:
            sku_detail.append({
                'salesman': d['salesman'], 'depo': d['depo'], 'bp': bp, 'dealer': d['dealer'],
                'sku_count': sku_count, 'bucket': bucket, 'sku_list': sorted(d['skus'])
            })
        dealer_detail.append({
            'salesman': d['salesman'], 'depo': d['depo'], 'bp': bp, 'dealer': d['dealer'],
            'device': d['Device'], 'macbook': d['Macbook'], 'acc': d['ACC'], 'total': total,
            'device_target': d['device_target'], 'macbook_target': d['macbook_target'], 'acc_target': d['acc_target'],
            'bo': bo, 'qvo': qvo, 'sku': sku_count,
            'sku_list': sorted(d['skus']),
            'device_items': sorted(
                d['device_items'].values(),
                key=lambda item: (-item['value'], item['type'].lower()),
            ),
            'is_target': d['is_target']
        })

    # Replace the depo-scoped BO subtotal with the salesman-owned BO total.
    # Revenue, QVO, SKU, and dealer detail intentionally remain depo-scoped.
    for salesman, actual_bo in bo_actual_by_salesman.items():
        if salesman not in salesmen:
            continue
        s = salesman_actual.setdefault(salesman, {
            'device':0.0,'macbook':0.0,'acc':0.0,'bo':0,'qvo':0,
            'sku_bins':{'1':0,'2-3':0,'4-6':0,'7-10':0,'>10':0}
        })
        s['bo'] = actual_bo

    all_people_found = set(salesman_target) | set(salesman_actual)
    all_people = [s for s in LOCKED_SALESMEN if s in all_people_found]
    table = []
    for salesman in all_people:
        a = salesman_actual.get(salesman, {'device':0,'macbook':0,'acc':0,'bo':0,'qvo':0,'sku_bins':{'1':0,'2-3':0,'4-6':0,'7-10':0,'>10':0}})
        t = salesman_target.get(salesman, {'device':0,'macbook':0,'acc':0,'bo':25,'qvo':0,'dealers':0})
        total = a['device'] + a['macbook'] + a['acc']
        target_total = t['device'] + t['macbook'] + t['acc']
        bo_target = t['bo'] or 25
        current_week = min(4, max(1, ((max([r.billing_date.day for r,owner in bo_billing if owner == salesman], default=1)-1)//7)+1))
        current_speed_target = weekly_targets(bo_target)[current_week]
        device_pct = a['device']/t['device']*100 if t['device'] else 0
        mac_pct = a['macbook']/t['macbook']*100 if t['macbook'] else 0
        acc_pct = a['acc']/t['acc']*100 if t['acc'] else 0
        bo_pct = a['bo']/bo_target*100 if bo_target else 0
        sales_pct = total/target_total*100 if target_total else 0
        bins = a['sku_bins']
        sku_score = min(bins['2-3'],13)+min(bins['4-6'],6)+min(bins['7-10'],5)+min(bins['>10'],2)
        sku_pct = sku_score/26*100
        health = 'Green' if a['bo'] >= current_speed_target else ('Amber' if a['bo'] >= max(1, math.ceil(current_speed_target*0.9)) else 'Red')
        table.append({
            'salesman':salesman,'device':a['device'],'macbook':a['macbook'],'acc':a['acc'],'total':total,
            'device_target':t['device'],'macbook_target':t['macbook'],'acc_target':t['acc'],'target_total':target_total,
            'device_pct':device_pct,'macbook_pct':mac_pct,'acc_pct':acc_pct,'sales_pct':sales_pct,
            'bo':a['bo'],'bo_target':bo_target,'bo_pct':bo_pct,'qvo':a['qvo'],'qvo_target':t['qvo'],
            'sku_bins':bins,'sku_score':sku_score,'sku_pct':sku_pct,'health':health,
            'overall_score':(bo_pct*0.45)+(sku_pct*0.30)+(min(sales_pct,150)*0.25)
        })

    leaderboard = sorted(table, key=lambda x:(x['overall_score'],x['bo'],x['qvo'],x['total']), reverse=True)
    for idx, x in enumerate(leaderboard,1): x['rank']=idx
    rank_map = {x['salesman']:x['rank'] for x in leaderboard}
    for x in table: x['rank']=rank_map.get(x['salesman'])
    table.sort(key=lambda x:x['rank'] or 999)

    # Speed Distribution per salesman and per week, cumulative BO. Future weeks are left blank.
    latest_in_scope = max([r.billing_date for r,_,_ in scoped_billing], default=None)

    # Time Gone follows the latest billing date in the active scope.
    working_days_elapsed, working_days_total, timegone_pct = working_day_progress(
        start, end, latest_in_scope or start
    )

    speed_rows = []
    for x in table:
        wk_targets = weekly_targets(x['bo_target'])
        weekly = {}
        # Speed Distribution follows the same salesman-owned BO rule and must
        # therefore include both mapped and UNMAPPED BP rows.
        person_rows = [(r, owner) for r, owner in bo_billing if owner == x['salesman']]
        for w in range(1,5):
            is_active = bool(latest_in_scope and latest_in_scope.day >= WEEK_START_DAY[w])
            if not is_active:
                weekly[w] = {'target':wk_targets[w], 'actual':None, 'pct':None, 'status':'Future'}
                continue
            cutoff_day = min(WEEK_END_DAY[w], end.day, latest_in_scope.day)
            cutoff = start.replace(day=cutoff_day)
            bo_by_bp = {}
            for r,owner in person_rows:
                if r.billing_date > cutoff:
                    continue
                bp = normalize_bp(r.sold_to_code)
                vals = bo_by_bp.setdefault(bp, {'device':0.0,'macbook':0.0})
                if r.category == 'Device': vals['device'] += float(r.nett_amount or 0)
                elif r.category == 'Macbook': vals['macbook'] += float(r.nett_amount or 0)
            actual = sum(1 for v in bo_by_bp.values() if v['device']+v['macbook']>0)
            pct = actual/wk_targets[w]*100 if wk_targets[w] else 0
            weekly[w] = {'target':wk_targets[w], 'actual':actual, 'pct':pct, 'status':'On Track' if actual>=wk_targets[w] else 'Need Push'}
        speed_rows.append({'salesman':x['salesman'],'weeks':weekly})

    sku_rows = [{'salesman':x['salesman'], **x['sku_bins']} for x in table]

    cards = {
        'sales': sum(x['total'] for x in table),
        'sales_target': sum(x['target_total'] for x in table),
        'device': sum(x['device'] for x in table), 'device_target':sum(x['device_target'] for x in table),
        'macbook': sum(x['macbook'] for x in table), 'macbook_target':sum(x['macbook_target'] for x in table),
        'acc': sum(x['acc'] for x in table), 'acc_target':sum(x['acc_target'] for x in table),
        'bo': sum(x['bo'] for x in table), 'bo_target':sum(x['bo_target'] for x in table),
        'qvo': sum(x['qvo'] for x in table), 'qvo_target':sum(x['qvo_target'] for x in table),
        'dealers': active_dealers, 'target_dealers':len(scoped_targets), 'salesmen':len(table)
    }
    cards['sales_pct'] = cards['sales']/cards['sales_target']*100 if cards['sales_target'] else 0
    cards['device_pct'] = cards['device']/cards['device_target']*100 if cards['device_target'] else 0
    cards['macbook_pct'] = cards['macbook']/cards['macbook_target']*100 if cards['macbook_target'] else 0
    cards['acc_pct'] = cards['acc']/cards['acc_target']*100 if cards['acc_target'] else 0
    cards['bo_rate'] = cards['bo']/cards['bo_target']*100 if cards['bo_target'] else 0
    cards['qvo_rate'] = cards['qvo']/cards['qvo_target']*100 if cards['qvo_target'] else 0
    # Achievement is sourced from Billing independently of Monthly Target.
    # Keep the target state explicit so templates can show Not yet instead of Rp0.
    target_available = bool(scoped_targets)

    uploads = UploadLog.query.order_by(UploadLog.uploaded_at.desc()).limit(6).all()
    target_uploads = TargetUploadLog.query.order_by(TargetUploadLog.uploaded_at.desc()).limit(6).all()
    return render_template(
        'dashboard.html', month=month,
        depo_filter=(depo_filters[0] if len(depo_filters) == 1 else ''),
        depo_filters=depo_filters, salesman_filters=salesman_filters,
        depos=depos, salesmen=salesmen, cards=cards, table=table, leaderboard=leaderboard,
        speed_rows=speed_rows, sku_rows=sku_rows, sku_detail=sku_detail, dealer_detail=dealer_detail,
        sku_targets=SKU_TARGETS, uploads=uploads, target_uploads=target_uploads,
        qvo_threshold=QVO_THRESHOLD, latest_in_scope=latest_in_scope,
        target_available=target_available,
        timegone_pct=timegone_pct,
        working_days_elapsed=working_days_elapsed,
        working_days_total=working_days_total
    )


@app.route('/program')
def program():
    """Program achievement based on Billing Detail Total Net Amount With Tax."""
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    depo_filter = normalize_text(request.args.get('depo'))
    salesman_filter = normalize_text(request.args.get('salesman'))
    latest = db.session.query(db.func.max(Billing.billing_date)).scalar()
    targets = MonthlyTarget.query.filter_by(month=month).all()
    program_bps = set(LOYALTY_CEMPAKA_BP)

    # Program tetap harus memiliki nama dealer dan salesman walaupun target bulan
    # yang dipilih baru berisi BP saja. Ambil metadata terbaru yang lengkap dari
    # Target Master, lalu gunakan Billing sebagai fallback terakhir.
    target_history = MonthlyTarget.query.order_by(
        MonthlyTarget.month.desc(), MonthlyTarget.id.desc()
    ).all()
    history_by_bp = {}
    for row in target_history:
        bp = normalize_bp(row.bp)
        if bp in program_bps:
            history_by_bp.setdefault(bp, []).append(row)

    billing_meta = {}
    # Do not filter the database using the raw BP value. Excel sometimes stores
    # the same BP as text, decimal-looking text, or a number. Normalise first in
    # Python so dealer and salesman metadata is found in every case.
    recent_billing = db.session.query(
        Billing.sold_to_code,
        Billing.sold_to_name,
        Billing.salesman,
        Billing.billing_date,
        Billing.id,
    ).order_by(Billing.billing_date.desc(), Billing.id.desc()).yield_per(1000)
    for row in recent_billing:
        bp = normalize_bp(row.sold_to_code)
        if bp in program_bps and bp not in billing_meta:
            billing_meta[bp] = row

    def program_metadata(bp):
        candidates = history_by_bp.get(bp, [])
        dealer = ''
        salesman = ''
        depo = ''
        for candidate in candidates:
            candidate_dealer = normalize_text(candidate.dealer)
            if not dealer and candidate_dealer and normalize_bp(candidate_dealer) != bp:
                dealer = candidate_dealer
            candidate_salesman = canonical_salesman(candidate.salesman)
            if not salesman and candidate_salesman:
                salesman = candidate_salesman
                depo = normalize_text(candidate.depo)
            if dealer and salesman:
                break
        billing_row = billing_meta.get(bp)
        if billing_row:
            if not dealer:
                dealer = normalize_text(billing_row.sold_to_name)
            if not salesman:
                salesman = canonical_salesman(billing_row.salesman)
        return {
            'dealer': dealer or bp,
            'salesman': salesman,
            'depo': depo or 'Cempaka',
        }

    metadata_by_bp = {bp: program_metadata(bp) for bp in program_bps}
    depos = sorted({meta['depo'] for meta in metadata_by_bp.values() if meta['depo']})
    available_salesmen = {meta['salesman'] for meta in metadata_by_bp.values() if meta['salesman']}
    salesmen = [name for name in LOCKED_SALESMEN if name in available_salesmen]
    start, end = month_range(month)
    achievement_by_bp = {}
    for raw_bp, total in db.session.query(
            Billing.sold_to_code,
            db.func.sum(Billing.nett_amount_with_tax),
        ).filter(
            Billing.billing_date >= start,
            Billing.billing_date <= end,
        ).group_by(Billing.sold_to_code).all():
        bp = normalize_bp(raw_bp)
        achievement_by_bp[bp] = achievement_by_bp.get(bp, 0.0) + float(total or 0)

    program_loyalty = []
    for bp, category in LOYALTY_CEMPAKA_BP.items():
        meta = metadata_by_bp[bp]
        depo = meta['depo']
        salesman = meta['salesman']
        if depo_filter and depo != depo_filter:
            continue
        if salesman_filter and salesman != salesman_filter:
            continue
        target_value = LOYALTY_TARGETS[category]
        achievement = achievement_by_bp.get(bp, 0.0)
        program_loyalty.append({
            'depo': depo,
            'bp': bp,
            'dealer': meta['dealer'],
            'category': category,
            'salesman': salesman,
            'target': target_value,
            'achievement': achievement,
            'pct': achievement / target_value * 100 if target_value else 0,
            'below_target': achievement < target_value,
        })
    program_loyalty.sort(key=lambda row: (-row['target'], row['dealer']))
    loyalty_target = sum(row['target'] for row in program_loyalty)
    loyalty_achievement = sum(row['achievement'] for row in program_loyalty)
    loyalty_totals = {
        'target': loyalty_target,
        'achievement': loyalty_achievement,
        'pct': loyalty_achievement / loyalty_target * 100 if loyalty_target else 0,
    }
    zero_totals = {'target': 0, 'achievement': 0, 'pct': 0}
    reward_notes = [
        {'category': 'CROWN', 'target': 5_000_000_000, 'reward_monthly': .01, 'reward_loyalty': .009},
        {'category': 'DIAMOND', 'target': 2_000_000_000, 'reward_monthly': .01, 'reward_loyalty': .006},
        {'category': 'GOLD', 'target': 500_000_000, 'reward_monthly': .0075, 'reward_loyalty': .0035},
        {'category': 'SILVER', 'target': 350_000_000, 'reward_monthly': .005, 'reward_loyalty': .0035},
        {'category': 'BRONZE', 'target': 50_000_000, 'reward_monthly': .005, 'reward_loyalty': None},
    ]
    return render_template(
        'program.html', month=month, latest=latest, depos=depos, salesmen=salesmen,
        depo_filter=depo_filter, salesman_filter=salesman_filter,
        program_loyalty=program_loyalty, program_ppg_paa=[], loyalty_totals=loyalty_totals,
        ppg_paa_totals=zero_totals, loyalty_reward_notes=reward_notes,
    )


PRICELIST_CATEGORIES = {
    'iphone': 'iPhone',
    'mac': 'Mac/MacBook',
    'ipad': 'iPad',
    'acc': 'ACC',
}


@app.route('/pricelist')
def pricelist():
    if 'user_id' not in session:
        session['username'] = 'Viewer'
        session['role'] = 'viewer'
    grouped = {key: [] for key in PRICELIST_CATEGORIES}
    rows = PricelistItem.query.order_by(PricelistItem.category, PricelistItem.sort_order, PricelistItem.id).all()
    for row in rows:
        if row.category in grouped:
            grouped[row.category].append(row)
    latest_uploads = {
        key: PricelistUploadLog.query.filter_by(category=key).order_by(PricelistUploadLog.uploaded_at.desc()).first()
        for key in PRICELIST_CATEGORIES
    }
    return render_template(
        'pricelist.html', grouped=grouped, category_labels=PRICELIST_CATEGORIES,
        latest_uploads=latest_uploads,
    )


@app.route('/pricelist/upload-image', methods=['POST'])
@admin_required
def upload_pricelist_image():
    payload = request.get_json(silent=True) or {}
    category = normalize_text(payload.get('category')).lower()
    filename = secure_filename(normalize_text(payload.get('filename')) or 'pricelist.jpg')
    rows = payload.get('rows') or []
    if category not in PRICELIST_CATEGORIES:
        return jsonify({'ok': False, 'message': 'Kategori pricelist tidak valid.'}), 400
    if not filename.lower().endswith(('.jpg', '.jpeg')):
        return jsonify({'ok': False, 'message': 'Format gambar harus JPG/JPEG.'}), 400
    if not isinstance(rows, list) or not rows:
        return jsonify({'ok': False, 'message': 'Tidak ada baris pricelist yang berhasil dibaca.'}), 400
    cleaned = []
    for index, raw in enumerate(rows):
        model = normalize_text(raw.get('model'))
        period = normalize_text(raw.get('period'))
        srp_promo = to_num(raw.get('srp_promo'), default=-1)
        stp_promo = to_num(raw.get('stp_promo'), default=-1)
        if not model or not period or srp_promo < 0 or stp_promo < 0:
            return jsonify({
                'ok': False,
                'message': f'Baris {index + 1} belum lengkap. Pastikan Model, SRP Promo, STP Promo, dan Period terbaca.',
            }), 400
        cleaned.append({
            'model': model, 'period': period,
            'srp_promo': srp_promo, 'stp_promo': stp_promo,
            'sort_order': index,
        })
    try:
        PricelistItem.query.filter_by(category=category).delete(synchronize_session=False)
        now = datetime.utcnow()
        for row in cleaned:
            db.session.add(PricelistItem(
                category=category, updated_at=now, updated_by=session.get('username'), **row,
            ))
        db.session.add(PricelistUploadLog(
            category=category, filename=filename, rows_loaded=len(cleaned),
            uploaded_by=session.get('username'), uploaded_at=now,
        ))
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Pricelist gagal disimpan: {exc}'}), 500
    return jsonify({
        'ok': True,
        'message': f'{len(cleaned)} baris {PRICELIST_CATEGORIES[category]} berhasil diperbarui.',
        'redirect': url_for('pricelist'),
    })


@app.route('/incentive')
def incentive():
    # Public Viewer may access SC incentive only.
    # Admin may access SC / ASH / TSH / LOB.
    if 'user_id' not in session:
        session['username'] = 'Viewer'
        session['role'] = 'viewer'

    latest = db.session.query(db.func.max(Billing.billing_date)).scalar()
    latest_target_month = db.session.query(db.func.max(MonthlyTarget.month)).scalar()
    default_month = (
        latest.strftime('%Y-%m')
        if latest else (latest_target_month or datetime.now().strftime('%Y-%m'))
    )

    month = request.args.get('month', default_month)
    requested_level = request.args.get('level', 'SC').upper()

    if session.get('role') == 'admin':
        level = requested_level if requested_level in INCENTIVE_ORG else 'SC'
        allowed_levels = ['SC', 'ASH', 'TSH', 'LOB']
        available_people = list(INCENTIVE_ORG[level].keys())
    else:
        # Viewer is restricted to SC incentive only for Cempaka + Serang + Cilegon.
        # Backend enforcement prevents access through URL manipulation.
        level = 'SC'
        allowed_levels = ['SC']
        viewer_sc = [
            'Zefanya Septania Simorangkir',
            'Rafhyski Alhasan',
            'Ikmah Novtianingrum',
        ]
        available_people = [
            name for name in viewer_sc
            if name in INCENTIVE_ORG['SC']
        ]

    person = request.args.get('name', '').strip()
    if person not in available_people:
        person = ''

    people_to_calculate = [person] if person else available_people
    rows = [
        calculate_incentive(level, name, month)
        for name in people_to_calculate
    ]

    # Compact summary totals by KPI.
    for r in rows:
        r['summary'] = {
            'Speed Distribution': 0,
            'SKU': 0,
            'QVO': 0,
            'Revenue': 0,
        }
        for d in r['detail']:
            if d['kpi'] == 'Speed Distribution':
                r['summary']['Speed Distribution'] += d['payout']
            elif d['kpi'] == 'SKU Penetration':
                r['summary']['SKU'] += d['payout']
            elif d['kpi'] == 'QVO':
                r['summary']['QVO'] += d['payout']
            elif d['kpi'] == 'Revenue':
                r['summary']['Revenue'] += d['payout']

    rows.sort(key=lambda x: (-x['earned'], x['person']))

    grand_earned = sum(x['earned'] for x in rows)
    grand_max = sum(x['max_incentive'] for x in rows)

    return render_template(
        'incentive.html',
        month=month,
        level=level,
        allowed_levels=allowed_levels,
        person=person,
        available_people=available_people,
        rows=rows,
        grand_earned=grand_earned,
        grand_max=grand_max,
        grand_pct=incentive_pct(grand_earned, grand_max),
    )


@app.route('/upload', methods=['POST'])
@admin_required
def upload():
    f = request.files.get('file')
    return_month = request.form.get('month', datetime.now().strftime('%Y-%m'))
    if not f or not f.filename:
        flash('Pilih file Billing Detail terlebih dahulu.', 'danger')
        return redirect(url_for('dashboard', month=return_month))
    if not f.filename.lower().endswith(('.xlsx','.xls')):
        flash('Format Billing Detail harus Excel (.xlsx/.xls).', 'danger')
        return redirect(url_for('dashboard', month=return_month))
    try:
        if not f.filename.lower().endswith('.xlsx'):
            raise ValueError('Upload ringan hanya mendukung format .xlsx. Simpan ulang file .xls sebagai .xlsx.')

        # Pass 1 only discovers the valid row count and replacement date range.
        # Nothing from the worksheet is retained in memory.
        rows_read = 0
        first_date = None
        last_date = None
        for rr in iter_billing_xlsx(f.stream):
            rows_read += 1
            billing_date = billing_date_value(rr['billing_date'])
            salesman_raw = normalize_text(rr['salesman'])
            code = normalize_bp(rr['sold_to_code'])
            if not billing_date or not salesman_raw or not code or salesman_raw.lower() == 'nan':
                continue
            first_date = billing_date if first_date is None else min(first_date, billing_date)
            last_date = billing_date if last_date is None else max(last_date, billing_date)

        if first_date is None or last_date is None:
            raise ValueError('Tidak ada baris Billing Detail valid yang dapat diimpor.')

        # Billing exports are snapshots. Replace the uploaded date range so old
        # with-tax values and records removed from the latest export cannot remain.
        replaced = Billing.query.filter(
            Billing.billing_date >= first_date,
            Billing.billing_date <= last_date,
        ).delete(synchronize_session=False)

        # Pass 2 parses and inserts in small batches. This keeps memory usage
        # nearly constant even when the Billing Detail contains many rows.
        seen_hashes = set()
        batch = []
        added = 0
        batch_size = 500
        for rr in iter_billing_xlsx(f.stream):
            billing_date = billing_date_value(rr['billing_date'])
            if not billing_date:
                continue
            salesman_raw = normalize_text(rr['salesman'])
            billing_document = normalize_text(rr['billing_document'])
            bill_item_no = normalize_text(rr['bill_item_no'])
            code_raw = normalize_text(rr['sold_to_code'])
            code = normalize_bp(code_raw)
            name = normalize_text(rr['sold_to_name'])
            group = normalize_text(rr['item_group'])
            article = normalize_text(rr['article'])
            qty = to_num(rr['quantity'])
            amount = to_num(rr['nett_amount'])
            amount_with_tax = to_num(rr['nett_amount_with_tax'])
            if not salesman_raw or not code or salesman_raw.lower() == 'nan':
                continue
            # Billing Document + Bill Item No uniquely identify a source line.
            # Without them, separate legitimate transactions with identical BP,
            # article, quantity and amount were incorrectly discarded as duplicates.
            h = row_hash([
                billing_document, bill_item_no, billing_date.isoformat(),
                salesman_raw, code_raw, name, group, article, qty, amount, amount_with_tax,
            ])
            if h in seen_hashes:
                continue
            seen_hashes.add(h)
            batch.append({
                'row_hash': h, 'billing_date': billing_date, 'salesman': salesman_raw,
                'sold_to_code': code, 'sold_to_name': name, 'item_group': group,
                'article': article, 'quantity': qty, 'nett_amount': amount,
                'nett_amount_with_tax': amount_with_tax,
                'category': classify(group), 'sku_key': sku_from_article(article, group),
            })
            if len(batch) >= batch_size:
                db.session.bulk_insert_mappings(Billing, batch)
                added += len(batch)
                batch.clear()
        if batch:
            db.session.bulk_insert_mappings(Billing, batch)
            added += len(batch)

        if not added:
            raise ValueError('Tidak ada baris Billing Detail valid yang dapat diimpor.')
        db.session.add(UploadLog(
            filename=secure_filename(f.filename), uploaded_by=session.get('username'),
            rows_read=rows_read, rows_added=added
        ))
        db.session.commit()
        flash(
            f'Billing berhasil: {added} baris No Tax disinkronkan untuk '
            f'{first_date.strftime("%d %b %Y")}–{last_date.strftime("%d %b %Y")}. '
            f'{replaced} baris lama diganti.',
            'success'
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Upload Billing gagal: {e}', 'danger')
    return redirect(url_for('dashboard', month=return_month))


@app.route('/upload-target', methods=['POST'])
@admin_required
def upload_target():
    f = request.files.get('target_file')
    target_month = request.form.get('target_month', datetime.now().strftime('%Y-%m'))
    if not f or not f.filename:
        flash('Pilih file Target Bulanan terlebih dahulu.', 'danger')
        return redirect(url_for('dashboard', month=target_month))
    if not f.filename.lower().endswith(('.xlsx','.xls')):
        flash('Format Target harus Excel (.xlsx/.xls).', 'danger')
        return redirect(url_for('dashboard', month=target_month))
    try:
        df = pd.read_excel(f, sheet_name=0)
        cmap = map_columns(df.columns, TARGET_ALIASES, 'Target Bulanan')
        collapsed = {}
        for _, rr in df.iterrows():
            bp = normalize_bp(rr[cmap['bp']])
            if not bp:
                continue
            salesman = canonical_salesman(rr[cmap['salesman']])
            depo = canonical_depo(rr[cmap['depo']])
            dealer_name = normalize_text(rr[cmap['dealer']])
            if not salesman or not dealer_name:
                continue
            rec = collapsed.setdefault(bp, {
                'depo':depo,'bp':bp,'dealer':dealer_name,'salesman':salesman,
                'device_target':0.0,'macbook_target':0.0,'acc_target':0.0,'bo_target':0,'qvo_target':0
            })
            rec['device_target'] += to_num(rr[cmap['device_target']])
            rec['macbook_target'] += to_num(rr[cmap['macbook_target']])
            rec['acc_target'] += to_num(rr[cmap['acc_target']])
            rec['bo_target'] += int(round(to_num(rr[cmap['bo_target']])))
            rec['qvo_target'] += int(round(to_num(rr[cmap['qvo_target']])))

        if not collapsed:
            raise ValueError('Tidak ada dealer/BP valid yang dapat dibaca.')

        # Monthly target is a snapshot: re-uploading a month cleanly replaces that month.
        MonthlyTarget.query.filter_by(month=target_month).delete(synchronize_session=False)
        for rec in collapsed.values():
            db.session.add(MonthlyTarget(month=target_month, **rec))
        db.session.add(TargetUploadLog(
            month=target_month, filename=secure_filename(f.filename), uploaded_by=session.get('username'),
            rows_read=len(df), dealers_loaded=len(collapsed)
        ))
        db.session.commit()
        flash(f'Target {target_month} berhasil: {len(collapsed)} dealer/BP dimuat.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Upload Target gagal: {e}', 'danger')
    return redirect(url_for('dashboard', month=target_month))


@app.route('/admin/targets')
@admin_required
def targets():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    rows = MonthlyTarget.query.filter_by(month=month).order_by(MonthlyTarget.depo, MonthlyTarget.salesman, MonthlyTarget.bp).all()
    summary = {}
    for t in rows:
        key = (t.depo,t.salesman)
        x = summary.setdefault(key, {'depo':t.depo,'salesman':t.salesman,'dealers':0,'device':0,'macbook':0,'acc':0,'bo':0,'qvo':0})
        x['dealers'] += 1; x['device'] += t.device_target or 0; x['macbook'] += t.macbook_target or 0; x['acc'] += t.acc_target or 0; x['bo'] += t.bo_target or 0; x['qvo'] += t.qvo_target or 0
    return render_template('targets.html', month=month, summary=sorted(summary.values(), key=lambda x:(x['depo'],x['salesman'])), target_count=len(rows))


@app.route('/admin/users', methods=['GET','POST'])
@admin_required
def users():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        role = request.form.get('role','viewer')
        if not username or not password:
            flash('Username dan password wajib diisi.', 'danger')
        elif User.query.filter_by(username=username).first():
            flash('Username sudah ada.', 'danger')
        else:
            db.session.add(User(username=username,password_hash=generate_password_hash(password),role=role))
            db.session.commit()
            flash('User berhasil dibuat.', 'success')
    return render_template('users.html', users=User.query.order_by(User.username).all())


@app.route('/stock/login', methods=['GET', 'POST'])
def stock_login():
    user_id = session.get('user_id')
    if user_id:
        user = db.session.get(User, user_id)
        if user and user.role == 'admin':
            return redirect(url_for('stock'))

    current_email = normalize_text(session.get('stock_email')).lower()
    if stock_session_is_valid():
        return redirect(url_for('stock'))
    if current_email:
        clear_stock_session()

    if request.method == 'POST':
        email = normalize_text(request.form.get('email')).lower()
        if email not in STOCK_ALLOWED_EMAILS:
            flash('Email ini tidak terdaftar untuk mengakses halaman Stock.', 'danger')
            return render_template('stock_login.html', email=email)

        now = datetime.utcnow()
        last_sent_raw = session.get('stock_otp_sent_at')
        if last_sent_raw:
            try:
                last_sent = datetime.fromisoformat(last_sent_raw)
                seconds_left = 60 - int((now - last_sent).total_seconds())
                if seconds_left > 0:
                    flash(f'Tunggu {seconds_left} detik sebelum meminta OTP baru.', 'danger')
                    return render_template('stock_login.html', email=email)
            except Exception:
                pass

        code = f'{secrets.randbelow(900000) + 100000:06d}'
        try:
            send_stock_otp(email, code)
        except Exception as exc:
            flash(f'OTP belum dapat dikirim: {exc}', 'danger')
            return render_template('stock_login.html', email=email)

        session['stock_otp_email'] = email
        session['stock_otp_hash'] = generate_password_hash(code)
        session['stock_otp_expires_at'] = (now + timedelta(minutes=10)).isoformat()
        session['stock_otp_sent_at'] = now.isoformat()
        session['stock_otp_attempts'] = 0
        flash('Kode OTP telah dikirim ke email Anda.', 'success')
        return redirect(url_for('stock_verify'))

    return render_template('stock_login.html', email='')


@app.route('/stock/verify', methods=['GET', 'POST'])
def stock_verify():
    email = normalize_text(session.get('stock_otp_email')).lower()
    otp_hash = session.get('stock_otp_hash')
    expires_raw = session.get('stock_otp_expires_at')
    if email not in STOCK_ALLOWED_EMAILS or not otp_hash or not expires_raw:
        return redirect(url_for('stock_login'))

    if request.method == 'POST':
        try:
            expires_at = datetime.fromisoformat(expires_raw)
        except Exception:
            expires_at = datetime.utcnow() - timedelta(seconds=1)
        if datetime.utcnow() > expires_at:
            flash('Kode OTP sudah kedaluwarsa. Silakan minta kode baru.', 'danger')
            return redirect(url_for('stock_login'))

        attempts = int(session.get('stock_otp_attempts', 0)) + 1
        session['stock_otp_attempts'] = attempts
        if attempts > 5:
            session.pop('stock_otp_hash', None)
            flash('Terlalu banyak percobaan. Silakan minta OTP baru.', 'danger')
            return redirect(url_for('stock_login'))

        code = re.sub(r'\D', '', request.form.get('code', ''))
        if len(code) != 6 or not check_password_hash(otp_hash, code):
            flash('Kode OTP tidak benar.', 'danger')
            return render_template('stock_verify.html', email=email)

        session.permanent = True
        session['stock_email'] = email
        session['stock_session_expires_at'] = (datetime.utcnow() + timedelta(minutes=STOCK_SESSION_MINUTES)).isoformat()
        for key in ('stock_otp_email', 'stock_otp_hash', 'stock_otp_expires_at', 'stock_otp_sent_at', 'stock_otp_attempts'):
            session.pop(key, None)
        return redirect(url_for('stock'))

    return render_template('stock_verify.html', email=email)


@app.route('/stock/logout')
def stock_logout():
    for key in ('stock_email', 'stock_session_expires_at', 'stock_otp_email', 'stock_otp_hash', 'stock_otp_expires_at', 'stock_otp_sent_at', 'stock_otp_attempts'):
        session.pop(key, None)
    return redirect(url_for('stock_login'))


@app.route('/stock')
@stock_required
def stock():
    snapshots = StockSnapshot.query.order_by(StockSnapshot.stock_date.desc()).all()
    requested_date = request.args.get('date', '').strip()
    selected = None
    if requested_date:
        try:
            selected = StockSnapshot.query.filter_by(stock_date=pd.to_datetime(requested_date).date()).first()
        except Exception:
            selected = None
    if selected is None and snapshots:
        selected = snapshots[0]

    cempaka_rows = []
    r5_rows = []
    totals = {short_name: 0.0 for short_name, _ in STOCK_DEPOTS}
    if selected:
        matrix = {}
        items = StockItem.query.filter_by(snapshot_id=selected.id).all()
        for item in items:
            key = (item.material, item.description)
            row = matrix.setdefault(key, {
                'material': item.material,
                'description': item.description,
                **{short_name: 0.0 for short_name, _ in STOCK_DEPOTS},
            })
            row[item.depot] += float(item.quantity or 0)
            totals[item.depot] += float(item.quantity or 0)

        r5_rows = sorted(matrix.values(), key=lambda row: (row['description'].casefold(), row['material']))
        for row in r5_rows:
            row['grand_total'] = sum(row[short_name] for short_name, _ in STOCK_DEPOTS)
        r5_rows = [row for row in r5_rows if row['grand_total'] != 0]
        cempaka_rows = [row for row in r5_rows if row['Cempaka'] != 0]

    totals['grand_total'] = sum(totals[short_name] for short_name, _ in STOCK_DEPOTS)
    return render_template(
        'stock.html',
        snapshots=snapshots,
        selected=selected,
        cempaka_rows=cempaka_rows,
        r5_rows=r5_rows,
        totals=totals,
        stock_email=stock_actor(),
        stock_is_admin=is_stock_admin(),
        today=datetime.now().strftime('%Y-%m-%d'),
    )


@app.route('/stock/upload', methods=['POST'])
@stock_required
@stock_admin_required
def stock_upload():
    file = request.files.get('stock_file')
    stock_date_raw = request.form.get('stock_date', '').strip()
    if not file or not file.filename:
        flash('Pilih file Excel stock terlebih dahulu.', 'danger')
        return redirect(url_for('stock'))
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('File stock harus berformat Excel (.xlsx atau .xls).', 'danger')
        return redirect(url_for('stock'))
    try:
        stock_date = pd.to_datetime(stock_date_raw).date()
        dataframe = pd.read_excel(file, sheet_name=0)
        snapshot, stored_rows = save_stock_snapshot(
            dataframe,
            stock_date,
            secure_filename(file.filename),
            stock_actor(),
            require_material_group=True,
        )
        flash(
            f'Stock {snapshot.stock_date.strftime("%d %b %Y")} berhasil: '
            f'{snapshot.app_rows} baris APP Storage Location 1001 diproses menjadi {stored_rows} posisi stock.',
            'success',
        )
    except Exception as exc:
        db.session.rollback()
        flash(f'Upload stock gagal: {exc}', 'danger')
    return redirect(url_for('stock', date=stock_date_raw))


@app.route('/stock/paste', methods=['POST'])
@stock_required
@stock_admin_required
def stock_paste():
    pasted = request.form.get('stock_paste', '').strip()
    stock_date_raw = request.form.get('stock_date', '').strip()
    if not pasted:
        flash('Paste data stock terlebih dahulu.', 'danger')
        return redirect(url_for('stock'))
    try:
        stock_date = pd.to_datetime(stock_date_raw).date()
        first_cells = [normalize_col(value) for value in pasted.splitlines()[0].split('\t')]
        has_header = 'material' in first_cells and any(value in first_cells for value in ('name 1', 'name1', 'depo', 'depot'))
        if has_header:
            dataframe = pd.read_csv(io.StringIO(pasted), sep='\t', dtype=str)
        else:
            dataframe = pd.read_csv(
                io.StringIO(pasted),
                sep='\t',
                dtype=str,
                header=None,
                names=['Material', 'Material Description', 'Name1', 'Unrestricted'],
            )
        snapshot, stored_rows = save_stock_snapshot(
            dataframe,
            stock_date,
            'Copy-paste',
            stock_actor(),
            require_material_group=False,
        )
        flash(
            f'Stock {snapshot.stock_date.strftime("%d %b %Y")} berhasil disimpan dari copy-paste: '
            f'{snapshot.app_rows} baris diproses menjadi {stored_rows} posisi stock.',
            'success',
        )
    except Exception as exc:
        db.session.rollback()
        flash(f'Paste stock gagal: {exc}', 'danger')
    return redirect(url_for('stock', date=stock_date_raw))


@app.route('/api/dealers')
@login_required
def api_dealers():
    # Retained for backward compatibility; V2 dashboard already receives dealer data server-side.
    return jsonify([])


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG') == '1')
